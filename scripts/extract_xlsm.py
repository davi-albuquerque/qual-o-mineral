"""Extract reference tables from `Qual o Minério 3.xlsm`.

Outputs:
    data/options/{brilho,traco,dureza,habito,luz,cor}.json   # {key, label}
    data/filter-index.json                                   # 4946 rows
    data/minerals-list.json                                  # 196 minerals from MINERAIS tab

Hardness is stored numerically in DADOS MINERAIS — bucketing happens in lib/hardness.ts.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _common import DATA_DIR, XLSM_FILE, first_token, slugify, write_json  # noqa: E402

OPTION_TABS = {
    "brilho": "BRILHO",
    "traco": "TRAÇO",
    "dureza": "DUREZA",
    "habito": "HÁBITO",
    "luz": "LUZ",
    "cor": "COR",
}


def label_to_key(label: str) -> str:
    """Stable ASCII key for each PT-BR label."""
    return slugify(label)


def extract_options(wb) -> dict:
    """Each option tab has one column of PT-BR labels."""
    out = {}
    for kind, tab_name in OPTION_TABS.items():
        ws = wb[tab_name]
        labels = []
        for row in ws.iter_rows(values_only=True, max_col=1):
            v = row[0]
            if v is None:
                continue
            label = str(v).strip()
            if not label:
                continue
            labels.append(label)
        # Deduplicate while preserving order
        seen = set()
        deduped = []
        for label in labels:
            if label in seen:
                continue
            seen.add(label)
            deduped.append({"key": label_to_key(label), "label": label})
        out[kind] = deduped
    return out


def extract_filter_index(wb) -> list[dict]:
    """Each row of DADOS MINERAIS becomes one FilterRow.

    Header: NOME DO MINERAL | TIPO | BRILHO | TRAÇO | DUREZA | DUREZA(empty) | HÁBITO | LUZ | COR
    Columns mapped to indices: 0, 1, 2, 3, 4, (5 ignored), 6, 7, 8.
    """
    ws = wb["DADOS MINERAIS"]
    rows = []
    skipped_blank = 0
    skipped_invalid = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_name = row[0]
        if raw_name is None or not str(raw_name).strip():
            skipped_blank += 1
            continue

        raw_name_str = str(raw_name).strip()
        name = first_token(raw_name_str)
        slug = slugify(name)
        type_ = (str(row[1]) if row[1] else "").strip()
        brilho = (str(row[2]) if row[2] else "").strip()
        traco = (str(row[3]) if row[3] else "").strip()
        dureza_raw = row[4]
        habito = (str(row[6]) if row[6] else "").strip()
        luz = (str(row[7]) if row[7] else "").strip()
        cor = (str(row[8]) if row[8] else "").strip()

        try:
            # Normalize PT-BR decimal comma to dot if it sneaks in as text.
            if isinstance(dureza_raw, str):
                dureza_num = float(dureza_raw.replace(",", "."))
            else:
                dureza_num = float(dureza_raw) if dureza_raw is not None else None
        except (TypeError, ValueError):
            skipped_invalid += 1
            continue
        if dureza_num is None:
            skipped_invalid += 1
            continue

        rows.append({
            "name": name,
            "rawName": raw_name_str,
            "slug": slug,
            "type": type_,
            "brilho": label_to_key(brilho) if brilho else "",
            "brilhoLabel": brilho,
            "traco": label_to_key(traco) if traco else "",
            "tracoLabel": traco,
            "durezaNum": dureza_num,
            "habito": label_to_key(habito) if habito else "",
            "habitoLabel": habito,
            "luz": label_to_key(luz) if luz else "",
            "luzLabel": luz,
            "cor": label_to_key(cor) if cor else "",
            "corLabel": cor,
        })

    print(f"  filter-index: {len(rows)} rows, skipped {skipped_blank} blank, {skipped_invalid} invalid hardness")
    return rows


def extract_minerals_list(wb) -> list[dict]:
    """The MINERAIS tab — one entry per canonical mineral (single column, name + formula)."""
    ws = wb["MINERAIS"]
    out = []
    seen = set()
    for row in ws.iter_rows(values_only=True, max_col=1):
        v = row[0]
        if v is None:
            continue
        full = str(v).strip()
        if not full:
            continue
        # Name = first whitespace-separated token (mirrors VBA macro)
        name = first_token(full)
        slug = slugify(name)
        # Heuristic: formula = remainder after the name + spaces, trimmed
        formula = full[len(name):].strip()
        if slug in seen:
            continue
        seen.add(slug)
        out.append({
            "slug": slug,
            "name": name,
            "rawName": full,
            "formula": formula,
        })
    return out


def main() -> None:
    print(f"Loading {XLSM_FILE.name} ...")
    wb = openpyxl.load_workbook(XLSM_FILE, data_only=True, read_only=True, keep_links=False)

    print("Extracting options ...")
    options = extract_options(wb)
    options_dir = DATA_DIR / "options"
    options_dir.mkdir(parents=True, exist_ok=True)
    for kind, opts in options.items():
        write_json(options_dir / f"{kind}.json", opts)
        print(f"  {kind}: {len(opts)} options")

    print("Extracting filter-index ...")
    rows = extract_filter_index(wb)
    write_json(DATA_DIR / "filter-index.json", rows)

    print("Extracting minerals list ...")
    minerals = extract_minerals_list(wb)
    write_json(DATA_DIR / "minerals-list.json", minerals)
    print(f"  minerals-list: {len(minerals)} unique minerals")

    print("Done.")


if __name__ == "__main__":
    main()
